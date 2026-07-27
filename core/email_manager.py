import os
import re
import threading
import openpyxl
from ui.colors import GREEN, RED, WHITE, EKL, LINE
from core.settings_manager import load_settings

file_lock = threading.Lock()

MAX_FORGET_EMAILS = 10000
EMAIL_FILE = "Email_List.txt"


# Extract email addresses from the best-matching column in an Excel file
def extract_from_excel(filename):
    try:
        wb = openpyxl.load_workbook(filename, data_only=True)
        sheet = wb.active
        target_col = None
        max_matches = 0

        # Scan first 20 rows to find the column with the most email-like values
        for col in range(1, sheet.max_column + 1):
            matches = 0
            for row in range(2, min(22, sheet.max_row + 1)):
                val = sheet.cell(row=row, column=col).value
                if val:
                    cleaned = str(val).strip()
                    if re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', cleaned):
                        matches += 1
            if matches > max_matches:
                max_matches = matches
                target_col = col

        if target_col is None:
            return None, "No email column found."

        emails = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=target_col, max_col=target_col, values_only=True):
            val = row[0]
            if val:
                cleaned = str(val).strip()
                if re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', cleaned):
                    emails.append(cleaned)
        return emails, None
    except Exception as e:
        return None, str(e)


# Load emails from Email_List.txt
def load_emails():
    with file_lock:
        try:
            with open(EMAIL_FILE, "r", encoding="utf-8", errors="ignore") as f:
                return [line.strip() for line in f if line.strip()]
        except FileNotFoundError:
            return []
        except Exception:
            return []


# Save emails to Email_List.txt
def save_emails(emails):
    with file_lock:
        with open(EMAIL_FILE, "w", encoding="utf-8") as f:
            for e in emails:
                f.write(e + "\n")


# Remove a processed email from Email_List.txt (thread-safe)
def remove_email(email):
    with file_lock:
        try:
            with open(EMAIL_FILE, "r", encoding="utf-8", errors="ignore") as f:
                lines = [line.strip() for line in f if line.strip()]
        except (FileNotFoundError, Exception):
            return

        if email in lines:
            lines.remove(email)

        try:
            with open(EMAIL_FILE, "w", encoding="utf-8") as f:
                for e in lines:
                    f.write(e + "\n")
        except Exception:
            pass


# Route file input based on settings (txt, multi-excel, or auto-detect)
def process_file_input():
    settings = load_settings()
    file_cfg = settings.get("file_input_settings", {})
    always_txt = file_cfg.get("always_use_txt", False)
    multi_excel = file_cfg.get("use_multiple_excel_files", False)

    if always_txt:
        return _load_txt()
    if multi_excel:
        return _load_multi_excel()
    return _load_auto()


# Load emails directly from Email_List.txt
def _load_txt():
    if not os.path.exists(EMAIL_FILE):
        print(f"{WHITE} '{EMAIL_FILE}' file was not found.")
        return []
    emails = load_emails()
    if not emails:
        print(f"{WHITE} '{EMAIL_FILE}' file is empty.")
        return []
    if len(emails) > MAX_FORGET_EMAILS:
        print(f"{RED} Too many emails! Maximum {MAX_FORGET_EMAILS} allowed.")
        print(f"{RED} You have {len(emails)} emails. Please reduce and try again.")
        return None
    print(f" {GREEN}[{RED}●{GREEN}] Selected File {EKL} {EMAIL_FILE}")
    return emails


# Load and merge emails from all Excel files in the directory
def _load_multi_excel():
    xlsx_files = [f for f in os.listdir('.') if f.endswith(".xlsx") and not f.startswith("~$")]
    if not xlsx_files:
        return _load_txt()

    print(f" {GREEN}[{RED}●{GREEN}] Found {len(xlsx_files)} Excel Files.")
    all_emails = []
    for f in xlsx_files:
        print(f"{WHITE} Extracting from {EKL} {f}...")
        ems, err = extract_from_excel(f)
        if ems:
            all_emails.extend(ems)
            print(f"{GREEN}  -> Found {len(ems)} emails.")
        else:
            print(f"{RED}  -> Failed: {err}")

    if not all_emails:
        print(f"{RED} No valid emails found in any Excel files.")
        return None

    all_emails = list(set(all_emails))

    if len(all_emails) > MAX_FORGET_EMAILS:
        print(f"{RED} Too many emails! Maximum {MAX_FORGET_EMAILS} allowed.")
        print(f"{RED} Total found: {len(all_emails)}. Please reduce files and try again.")
        return None

    save_emails(all_emails)
    print(f"\n {GREEN}[{RED}●{GREEN}] Total Unique Emails {EKL} {len(all_emails)}")
    print(f" {GREEN}[{RED}●{GREEN}] Saved to '{EMAIL_FILE}'\n")
    return all_emails


# Auto-detect input source: use Excel if available, fall back to txt
def _load_auto():
    xlsx_files = [f for f in os.listdir('.') if f.endswith(".xlsx") and not f.startswith("~$")]
    if not xlsx_files:
        return _load_txt()

    filename = None
    if len(xlsx_files) == 1:
        filename = xlsx_files[0]
    else:
        print(f" {GREEN}[{RED}●{GREEN}] Found {len(xlsx_files)} Excel Files:")
        for idx, f in enumerate(xlsx_files, 1):
            print(f" {GREEN}[{RED}{idx}{GREEN}] {f}")
        print(f"{LINE}")
        while True:
            try:
                choice = input(f" {GREEN}[{RED}●{GREEN}] Select File (1-{len(xlsx_files)}) {EKL} ").strip()
                if choice.isdigit():
                    idx = int(choice) - 1
                    if 0 <= idx < len(xlsx_files):
                        filename = xlsx_files[idx]
                        break
                print(f"{RED} Invalid selection!")
            except KeyboardInterrupt:
                raise
            except Exception:
                pass

    print(f" {GREEN}[{RED}●{GREEN}] Selected File {EKL} {filename}\n")
    ems, err = extract_from_excel(filename)

    if ems:
        if len(ems) > MAX_FORGET_EMAILS:
            print(f"{RED} Too many emails! Maximum {MAX_FORGET_EMAILS} allowed.")
            print(f"{RED} Found {len(ems)} emails in {filename}. Please reduce and try again.")
            return None
        save_emails(ems)
        print(f" {GREEN}[{RED}●{GREEN}] Extracted {len(ems)} emails from {filename}")
        print(f" {GREEN}[{RED}●{GREEN}] Saved to '{EMAIL_FILE}'\n")
        return ems
    else:
        print(f"{RED} Error: {err}")
        return None
